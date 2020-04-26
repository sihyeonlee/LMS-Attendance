print("Prepare to Load Excel Library...")

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Color
from getpass import getpass

wb = openpyxl.Workbook()
ws = wb.active
ws.column_dimensions['A'].width = 48
ws.column_dimensions['C'].width = 82
ws.column_dimensions['D'].width = 9

ws['A3'] = "과목명"
ws['B3'] = "주차"
ws['C3'] = "강의 제목"
ws['D3'] = "출석률"

excel_index = 4

print("Username: ")
username = str(input())
password = str(getpass())

# Site Information Start
url = 'https://myclass.ssu.ac.kr/login/index.php'

header = {
    'Referer': 'https://myclass.ssu.ac.kr/login.php',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:75.0) Gecko/20100101 Firefox/75.0'
}

data = {
    'username': username,
    'password': password
}
# Site Information End

# Session Make
session = requests.session()
session.post(url, headers=header, data=data)                            # Login


main_site = session.get("http://myclass.ssu.ac.kr")

parser = BeautifulSoup(main_site.content, "html.parser")

course = parser.find_all("div", class_='course_box')                    # Course List Get

if len(course) == 0:
    print("로그인 실패")
    print("아이디나 비밀번호 확인 후 다시 시도해주세요.")
    print("3회 이상 실패시 사이트에서 인증 잠김 확인 해주세요.")
    input()
    exit()
else:
    print("로그인 성공")

course_id_list = []
course_title_list = []

for i in course:
    course_link = i.find_all("a", class_='course_link')[0]['href']      # Course Link
    course_title = i.find("h3").text                                    # Course Title
    course_id = course_link[-5:]
    course_id_list.append(course_id)
    course_title_list.append(course_title)

attend_link = "http://myclass.ssu.ac.kr/report/ubcompletion/user_progress.php?id="

for index, i in enumerate(course_id_list):
    # print(course_title_list[index])
    print("총 과목 : %d 중 %d 과목 처리 중..." % (len(course_id_list), index + 1))
    course_excel = 'A' + str(excel_index)
    ws[course_excel] = course_title_list[index]

    fix_attend_link = attend_link + str(i)
    # print(fix_attend_link)

    attend_site = session.get(fix_attend_link)

    if attend_site.ok is True:
        pass

    attend_parser = BeautifulSoup(attend_site.content, "html.parser")
    raw_table = attend_parser.find("table", class_="table table-bordered user_progress")
    try:
        table_trs = raw_table.find_all('tr')
    except:
        print("숭실사이버대학교 학점 교류 과목은 지원하지 않습니다.")
        title_index += 1
        continue

    table_trs = table_trs[1:]

    for index_, week in enumerate(table_trs):
        if week.find("div", class_="sectiontitle") is not None:
            raw_data = week.find_all("td", class_="text-center")
            left_data = week.find_all("td", class_="text-left")

            if len(raw_data) >= 4:
                week_data = raw_data.pop(0)
                # print(week_data.text, "주차")
                week_excel = 'B' + str(excel_index)
                ws[week_excel] = week_data.text
                title_index = 0

                while len(raw_data) != 0:
                    try:
                        title = left_data[title_index].text

                    except:
                        del raw_data[0:3]
                        title_index += 1
                        excel_index += 1
                        continue

                    # print(title)
                    # print(raw_data[2].text)

                    try:
                        percent = float(raw_data[2].text[:len(raw_data[2].text)-1])

                    except:
                        percent = -1

                    title_excel = 'C' + str(excel_index)
                    ws[title_excel] = title
                    percent_excel = 'D' + str(excel_index)
                    ws[percent_excel] = raw_data[2].text
                    if percent == -1:
                        pass
                    elif percent > 95:
                        ca = ws[percent_excel]
                        ca.fill = PatternFill(patternType='solid', fgColor=Color('45F542'))
                    elif percent > 85:
                        ca = ws[percent_excel]
                        ca.fill = PatternFill(patternType='solid', fgColor=Color('F5EF42'))
                    else:
                        ca = ws[percent_excel]
                        ca.fill = PatternFill(patternType='solid', fgColor=Color('F54242'))

                    del raw_data[0:3]
                    title_index += 1
                    excel_index += 1
    excel_index += 1

while True:
    try:
        wb.save("스마트캠퍼스 출석표.xlsx")
        print("엑셀 파일 저장 완료")
        input()
        break

    except:
        print("엑셀 파일이 열려 있는 경우 저장이 불가능합니다.")
        print("엑셀 파일을 종료 후 아무 키나 눌러주세요.")
        input()